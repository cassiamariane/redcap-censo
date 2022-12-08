import openpyxl

tabela = openpyxl.load_workbook('arquivorequerido.xlsx')

aba = tabela.active

print("Como a tabela é lida:")

for cols in aba.iter_cols():
  for row in cols:
    print(row.value)

def busca(dado):
  i = 0
  j = 0
  linha = 0
  coluna = 0

  for cols in aba.iter_cols():
      i += 1
      j = 0
      for row in cols:
        j += 1
        if(row.value == dado):
          coluna = i
          linha = j 
          while(aba.cell(linha, coluna).value != "None"):
            linha +=1

  return aba.cell(linha, coluna).offset(1,0).value

CUSTOMERID = busca('CUSTOMERID')
FIRSTNAME = busca('FIRSTNAME')
LASTNAME = busca('LASTNAME')
SEX = busca('SEX')
HEIGHT = busca('HEIGHT')
WEIGHT = busca('WEIGHT')
BMR = busca('BMR')
FATP = busca('FATP')
FATM = busca('FATM')
FFM = busca('FFM')
TBW = busca('TBW')
PMM = busca('PMM')
IMP = busca('IMP')
BMI = busca('BMI')
VFATL = busca('VFATL')
BONEM = busca('BONEM')
ECW = busca('ECW')
ICW = busca('ICW')
METAAGE = busca('METAAGE')
PHASEANGLE = busca('PHASEANGLE')
DCI = busca('DCI')
PAL = busca('PAL')
CLOTHES = busca('CLOTHES')
PHYSRATE = busca('PHYSRATE')
RLFATP = busca('RLFATP')
RLFATM = busca('RLFATM')
RLFFM = busca('RLFFM')
RLPMM = busca('RLPMM')
RLIMP = busca('RLIMP')
LLFATP = busca('LLFATP')
LLFATM = busca('LLFATM')
LLFFM = busca('LLFFM')
LLPMM = busca('LLPMM')
LLIMP = busca('LLIMP')
RAFATP = busca('RAFATP')
RAFATM = busca('RAFATM') 
RAFFM = busca('RAFFM')
RAPMM = busca('RAPMM')
RAIMP = busca('RAIMP')
LAFATP = busca('LAFATP')
LAFATM = busca('LAFATM')
LAFFM = busca('LAFFM')
LAPMM = busca('LAPMM')
LAIMP = busca('LAIMP')
TRFATP = busca('TRFATP')
TRFATM = busca('TRFATM')
TRFFM = busca('TRFFM')
TRPMM = busca('TRPMM')


dados = [CUSTOMERID, FIRSTNAME, LASTNAME, SEX, HEIGHT, WEIGHT, BMR, FATP, FATM, FFM, TBW, PMM, IMP, BMI, VFATL, BONEM, ECW, ICW, METAAGE, PHASEANGLE, DCI, PAL, CLOTHES, PHYSRATE, RLFATP, RLFATM, RLFFM, RLPMM, RLIMP, LLFATP, LLFATM, LLFFM, LLPMM, LLIMP, RAFATP, RAFATM, RAFFM, RAPMM, RAIMP, LAFATP, LAFATM, LAFFM, LAPMM, LAIMP, TRFATP, TRFATM, TRFFM, TRPMM]
print("Dados obtidos:")
print(dados)

#tabelaFinal = openpyxl.load_workbook('principal.xlsx')
#abaTF = tabelaFinal.active
#tabelaFinal.save('principal.xlsx')
